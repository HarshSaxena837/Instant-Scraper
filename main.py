from bs4 import BeautifulSoup
# ('span', class_="css-1n7hynb")
import requests
import openpyxl
excel = openpyxl.Workbook()
# print(excel.sheetnames)
sheet = excel.active
sheet.title = 'NewYork Times Contributors List'
# print(excel.sheetnames)
sheet.append(['Contributors Name'])
try:
  #  <div class="css-1fbiiks e1j8vip06"><h1 class="css-tp63b9 e16wpn5v0">Shira Ovide</h1></div>
    #stream-panel > div.css-13mho3u
    # <span class="css-1n7hynb">
    #stream-panel > div.css-13mho3u > ol > li:nth-child(1) > div > div.css-1l4spti > a > div.css-1i4y2t3.e140qd2t0 > p > span
    source = requests.get('https://www.nytimes.com/section/technology')
    
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    names = soup.find('div', class_="css-13mho3u").find_all('li')
    for name in names:
      # contributors = name.find('p', class_="css-g0iztv").text
      contributors = name.find('span', class_="css-1n7hynb").text
      print(contributors)
      sheet.append([contributors])
    # print(len(names))

except Exception as e:
  print(e)

excel.save('NewYork Times Contributors List.xlsx')